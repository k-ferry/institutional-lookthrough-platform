"""Holdings API endpoints — paginated list, filter options, sources, and exports."""

import csv
import io
from collections import defaultdict
from datetime import date
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func
from sqlalchemy.orm import Session

from src.lookthrough.auth.dependencies import get_current_user, get_db
from src.lookthrough.db.models import (
    DimCompany,
    DimFund,
    EntityResolutionLog,
    FactFundReport,
    FactReportedHolding,
    User,
)

router = APIRouter(prefix="/api/holdings", tags=["holdings"])


# ---------------------------------------------------------------------------
# Shared column expressions
# ---------------------------------------------------------------------------

def _company_name_col():
    return func.coalesce(DimCompany.company_name, FactReportedHolding.raw_company_name)

def _sector_col():
    return func.coalesce(DimCompany.primary_sector, "Unclassified")

def _industry_col():
    return func.coalesce(DimCompany.primary_industry, "Unclassified")

def _country_col():
    # Prefer the holding's reported country; fall back to canonical, then Unknown
    return func.coalesce(
        FactReportedHolding.reported_country,
        DimCompany.primary_country,
        "Unknown",
    )


# ---------------------------------------------------------------------------
# GET /api/holdings
# ---------------------------------------------------------------------------


@router.get("")
def list_holdings(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    search: Optional[str] = Query(None),
    fund_id: Optional[str] = Query(None),
    sector: Optional[str] = Query(None),
    industry: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
    as_of_date: Optional[str] = Query(None),
    has_value: Optional[bool] = Query(None),
    sort_by: str = Query("reported_value_usd"),
    sort_dir: str = Query("desc"),
    db: Session = Depends(get_db),
    _current_user: User = Depends(get_current_user),
) -> dict:
    """Return a paginated list of holdings with full metadata and optional filters."""
    # Subquery: total reported value per fund-report period (for pct_of_fund)
    fr_totals_sq = (
        db.query(
            FactReportedHolding.fund_report_id.label("frid"),
            func.sum(FactReportedHolding.reported_value_usd).label("report_total"),
        )
        .group_by(FactReportedHolding.fund_report_id)
        .subquery()
    )

    company_name_col = _company_name_col()
    sector_col = _sector_col()
    industry_col = _industry_col()
    country_col = _country_col()
    pct_col = (
        FactReportedHolding.reported_value_usd
        * 100.0
        / func.nullif(fr_totals_sq.c.report_total, 0)
    ).label("pct_of_fund")

    query = (
        db.query(
            FactReportedHolding.reported_holding_id,
            FactReportedHolding.company_id,
            company_name_col.label("company_name"),
            DimFund.fund_id,
            DimFund.fund_name,
            sector_col.label("sector"),
            industry_col.label("industry"),
            country_col.label("country"),
            FactReportedHolding.reported_value_usd,
            pct_col,
            FactReportedHolding.as_of_date,
            FactReportedHolding.source,
            FactReportedHolding.cost_basis_usd,
            FactReportedHolding.ownership_pct,
        )
        .join(FactFundReport, FactReportedHolding.fund_report_id == FactFundReport.fund_report_id)
        .join(DimFund, FactFundReport.fund_id == DimFund.fund_id)
        .outerjoin(DimCompany, FactReportedHolding.company_id == DimCompany.company_id)
        .outerjoin(fr_totals_sq, fr_totals_sq.c.frid == FactReportedHolding.fund_report_id)
    )

    if search:
        query = query.filter(company_name_col.ilike(f"%{search}%"))
    if fund_id:
        query = query.filter(DimFund.fund_id == fund_id)
    if sector:
        query = query.filter(sector_col == sector)
    if industry:
        query = query.filter(industry_col == industry)
    if country:
        query = query.filter(country_col == country)
    if source:
        query = query.filter(FactReportedHolding.source == source)
    if as_of_date:
        query = query.filter(FactReportedHolding.as_of_date == as_of_date)
    if has_value:
        query = query.filter(FactReportedHolding.reported_value_usd.isnot(None))

    sort_col_map = {
        "reported_value_usd": FactReportedHolding.reported_value_usd,
        "company_name": company_name_col,
        "fund_name": DimFund.fund_name,
        "sector": sector_col,
        "industry": industry_col,
        "country": country_col,
        "as_of_date": FactReportedHolding.as_of_date,
    }
    col = sort_col_map.get(sort_by, FactReportedHolding.reported_value_usd)
    query = query.order_by(
        col.asc().nullslast() if sort_dir == "asc" else col.desc().nullslast()
    )

    total = query.count()
    total_pages = max(1, (total + page_size - 1) // page_size)
    rows = query.offset((page - 1) * page_size).limit(page_size).all()

    items = [
        {
            "holding_id": row.reported_holding_id,
            "company_id": row.company_id,
            "company_name": row.company_name,
            "fund_id": row.fund_id,
            "fund_name": row.fund_name,
            "sector": row.sector,
            "industry": row.industry,
            "country": row.country,
            "reported_value_usd": (
                float(row.reported_value_usd) if row.reported_value_usd is not None else None
            ),
            "pct_of_fund": (
                round(float(row.pct_of_fund), 2) if row.pct_of_fund is not None else None
            ),
            "as_of_date": row.as_of_date,
            "source": row.source,
            "cost_basis_usd": (
                float(row.cost_basis_usd) if row.cost_basis_usd is not None else None
            ),
            "ownership_pct": (
                float(row.ownership_pct) if row.ownership_pct is not None else None
            ),
        }
        for row in rows
    ]

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
    }


# ---------------------------------------------------------------------------
# GET /api/holdings/sources
# ---------------------------------------------------------------------------


@router.get("/sources")
def get_sources(
    db: Session = Depends(get_db),
    _current_user: User = Depends(get_current_user),
) -> list:
    """Return distinct sources with holding counts and latest date."""
    rows = (
        db.query(
            FactReportedHolding.source,
            func.count(FactReportedHolding.reported_holding_id).label("count"),
            func.max(FactReportedHolding.as_of_date).label("latest_as_of_date"),
        )
        .filter(FactReportedHolding.source.isnot(None))
        .group_by(FactReportedHolding.source)
        .order_by(func.count(FactReportedHolding.reported_holding_id).desc())
        .all()
    )
    return [
        {
            "source": row.source,
            "count": row.count,
            "latest_as_of_date": row.latest_as_of_date,
        }
        for row in rows
    ]


# ---------------------------------------------------------------------------
# GET /api/holdings/filters
# ---------------------------------------------------------------------------


@router.get("/filters")
def get_filter_options(
    db: Session = Depends(get_db),
    _current_user: User = Depends(get_current_user),
) -> dict:
    """Return available filter options, global totals, and source breakdown."""
    company_name_col = _company_name_col()
    sector_col = _sector_col()
    industry_col = _industry_col()
    country_col = _country_col()

    # Global totals (unfiltered)
    totals = db.query(
        func.count(FactReportedHolding.reported_holding_id).label("total_holdings"),
        func.sum(FactReportedHolding.reported_value_usd).label("total_exposure"),
    ).one()

    # Funds that have holdings
    fund_rows = (
        db.query(DimFund.fund_id, DimFund.fund_name, DimFund.source.label("fund_source"))
        .join(FactFundReport, FactFundReport.fund_id == DimFund.fund_id)
        .join(FactReportedHolding, FactReportedHolding.fund_report_id == FactFundReport.fund_report_id)
        .distinct()
        .order_by(DimFund.fund_name)
        .all()
    )
    funds = [{"id": r.fund_id, "name": r.fund_name, "source": r.fund_source} for r in fund_rows]

    # Distinct coalesced sectors
    sector_rows = (
        db.query(sector_col.label("sector"))
        .select_from(FactReportedHolding)
        .outerjoin(DimCompany, FactReportedHolding.company_id == DimCompany.company_id)
        .distinct()
        .order_by(sector_col)
        .all()
    )
    sectors = [r.sector for r in sector_rows if r.sector]

    # Distinct industries with parent sector (for cascade filtering)
    industry_rows = (
        db.query(
            industry_col.label("industry"),
            sector_col.label("sector"),
        )
        .select_from(FactReportedHolding)
        .outerjoin(DimCompany, FactReportedHolding.company_id == DimCompany.company_id)
        .distinct()
        .order_by(industry_col)
        .all()
    )
    industries = [
        {"name": r.industry, "sector": r.sector}
        for r in industry_rows
        if r.industry
    ]

    # Distinct countries
    country_rows = (
        db.query(country_col.label("country"))
        .select_from(FactReportedHolding)
        .outerjoin(DimCompany, FactReportedHolding.company_id == DimCompany.company_id)
        .distinct()
        .order_by(country_col)
        .all()
    )
    countries = [r.country for r in country_rows if r.country and r.country != "Unknown"]
    if any(r.country == "Unknown" for r in country_rows):
        countries.append("Unknown")

    # Distinct as_of_dates (descending — most recent first)
    date_rows = (
        db.query(FactReportedHolding.as_of_date)
        .filter(FactReportedHolding.as_of_date.isnot(None))
        .distinct()
        .order_by(FactReportedHolding.as_of_date.desc())
        .all()
    )
    dates = [r.as_of_date for r in date_rows]

    # Distinct sources
    source_rows = (
        db.query(FactReportedHolding.source)
        .filter(FactReportedHolding.source.isnot(None))
        .distinct()
        .order_by(FactReportedHolding.source)
        .all()
    )
    sources = [r.source for r in source_rows]

    # Count of holdings with a reported value
    has_value_count = (
        db.query(func.count(FactReportedHolding.reported_holding_id))
        .filter(FactReportedHolding.reported_value_usd.isnot(None))
        .scalar()
        or 0
    )

    return {
        "total_holdings": totals.total_holdings or 0,
        "total_exposure": float(totals.total_exposure) if totals.total_exposure else 0.0,
        "funds": funds,
        "sectors": sectors,
        "industries": industries,
        "countries": countries,
        "dates": dates,
        "sources": sources,
        "has_value_count": has_value_count,
    }


# ---------------------------------------------------------------------------
# Export helpers
# ---------------------------------------------------------------------------

EXPORT_COLUMNS = [
    "Company Name",
    "Fund",
    "Sector",
    "Industry",
    "Country",
    "Reported Value (USD)",
    "% of Fund",
    "As of Date",
    "Source",
    "Extraction Confidence",
    "Match Method",
    "Match Confidence",
]

_HEADER_FILL = PatternFill("solid", fgColor="0F2B5B")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _build_export_query(
    db: Session,
    search: Optional[str],
    fund_id: Optional[str],
    sector: Optional[str],
    industry: Optional[str],
    country: Optional[str],
    source: Optional[str],
    as_of_date: Optional[str],
    has_value: Optional[bool],
):
    """Un-paginated query for exports — all filters, plus entity resolution."""
    company_name_col = _company_name_col()
    sector_col = _sector_col()
    industry_col = _industry_col()
    country_col = _country_col()

    query = (
        db.query(
            company_name_col.label("company_name"),
            DimFund.fund_id,
            DimFund.fund_name,
            sector_col.label("sector"),
            industry_col.label("industry"),
            country_col.label("country"),
            FactReportedHolding.reported_value_usd,
            FactReportedHolding.extraction_confidence,
            FactReportedHolding.as_of_date,
            FactReportedHolding.source,
            EntityResolutionLog.match_method,
            EntityResolutionLog.match_confidence,
        )
        .join(FactFundReport, FactReportedHolding.fund_report_id == FactFundReport.fund_report_id)
        .join(DimFund, FactFundReport.fund_id == DimFund.fund_id)
        .outerjoin(DimCompany, FactReportedHolding.company_id == DimCompany.company_id)
        .outerjoin(
            EntityResolutionLog,
            FactReportedHolding.reported_holding_id == EntityResolutionLog.reported_holding_id,
        )
    )

    if search:
        query = query.filter(company_name_col.ilike(f"%{search}%"))
    if fund_id:
        query = query.filter(DimFund.fund_id == fund_id)
    if sector:
        query = query.filter(sector_col == sector)
    if industry:
        query = query.filter(industry_col == industry)
    if country:
        query = query.filter(country_col == country)
    if source:
        query = query.filter(FactReportedHolding.source == source)
    if as_of_date:
        query = query.filter(FactReportedHolding.as_of_date == as_of_date)
    if has_value:
        query = query.filter(FactReportedHolding.reported_value_usd.isnot(None))

    return query


def _rows_to_dicts(rows) -> list[dict]:
    """Convert query rows to export dicts. % of Fund uses per-fund-per-date totals."""
    # Compute totals per (fund_id, as_of_date) for accurate pct
    period_totals: dict[tuple, float] = defaultdict(float)
    for row in rows:
        if row.reported_value_usd is not None:
            period_totals[(row.fund_id, row.as_of_date)] += float(row.reported_value_usd)

    result = []
    for row in rows:
        value = float(row.reported_value_usd) if row.reported_value_usd is not None else None
        pt = period_totals.get((row.fund_id, row.as_of_date), 0.0)
        pct = round(value / pt * 100, 4) if (value is not None and pt > 0) else None
        result.append({
            "Company Name": row.company_name or "",
            "Fund": row.fund_name or "",
            "Sector": row.sector or "",
            "Industry": row.industry or "",
            "Country": row.country or "",
            "Reported Value (USD)": value if value is not None else "",
            "% of Fund": pct if pct is not None else "",
            "As of Date": row.as_of_date or "",
            "Source": row.source or "",
            "Extraction Confidence": (
                float(row.extraction_confidence) if row.extraction_confidence is not None else ""
            ),
            "Match Method": row.match_method or "",
            "Match Confidence": (
                float(row.match_confidence) if row.match_confidence is not None else ""
            ),
        })
    return result


# ---------------------------------------------------------------------------
# GET /api/holdings/export — CSV
# ---------------------------------------------------------------------------


@router.get("/export")
def export_holdings_csv(
    search: Optional[str] = Query(None),
    fund_id: Optional[str] = Query(None),
    sector: Optional[str] = Query(None),
    industry: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
    as_of_date: Optional[str] = Query(None),
    has_value: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """Export all holdings matching current filters as a CSV download."""
    rows = _build_export_query(
        db, search, fund_id, sector, industry, country, source, as_of_date, has_value
    ).all()
    dicts = _rows_to_dicts(rows)

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=EXPORT_COLUMNS)
    writer.writeheader()
    writer.writerows(dicts)

    filename = f"lookthrough_holdings_{date.today().isoformat()}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# GET /api/holdings/export/excel — XLSX
# ---------------------------------------------------------------------------


@router.get("/export/excel")
def export_holdings_excel(
    search: Optional[str] = Query(None),
    fund_id: Optional[str] = Query(None),
    sector: Optional[str] = Query(None),
    industry: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
    as_of_date: Optional[str] = Query(None),
    has_value: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """Export all holdings matching current filters as an Excel (.xlsx) download."""
    rows = _build_export_query(
        db, search, fund_id, sector, industry, country, source, as_of_date, has_value
    ).all()
    dicts = _rows_to_dicts(rows)

    wb = openpyxl.Workbook()

    # --- Sheet 1: Holdings ---
    ws = wb.active
    ws.title = "Holdings"
    ws.append(EXPORT_COLUMNS)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_dict in dicts:
        ws.append([row_dict[col] for col in EXPORT_COLUMNS])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet("Summary")

    def _styled_header(sheet, values):
        sheet.append(values)
        for cell in sheet[sheet.max_row]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT

    _NA = "N/A - value data not available for all holdings"
    total_aum = sum(
        float(r.reported_value_usd) for r in rows if r.reported_value_usd is not None
    )
    _styled_header(ws2, ["Metric", "Value"])
    ws2.append(["Total Holdings", len(rows)])
    ws2.append(["Total AUM (USD)", round(total_aum, 2) if total_aum > 0 else _NA])
    ws2.append(["Export Date", date.today().isoformat()])
    ws2.append([])

    fund_agg: dict[str, dict] = defaultdict(lambda: {"count": 0, "aum": 0.0, "has_value": False})
    for row in rows:
        fund_agg[row.fund_name]["count"] += 1
        if row.reported_value_usd is not None:
            fund_agg[row.fund_name]["aum"] += float(row.reported_value_usd)
            fund_agg[row.fund_name]["has_value"] = True
    _styled_header(ws2, ["Fund", "Holdings", "AUM (USD)"])
    for name, stats in sorted(fund_agg.items(), key=lambda x: x[1]["aum"], reverse=True):
        aum_cell = round(stats["aum"], 2) if stats["has_value"] else _NA
        ws2.append([name, stats["count"], aum_cell])
    ws2.append([])

    sector_agg: dict[str, dict] = defaultdict(lambda: {"count": 0, "aum": 0.0, "has_value": False})
    for row in rows:
        s = row.sector or "Unclassified"
        sector_agg[s]["count"] += 1
        if row.reported_value_usd is not None:
            sector_agg[s]["aum"] += float(row.reported_value_usd)
            sector_agg[s]["has_value"] = True
    _styled_header(ws2, ["Sector", "Holdings", "AUM (USD)"])
    for name, stats in sorted(sector_agg.items(), key=lambda x: x[1]["aum"], reverse=True):
        aum_cell = round(stats["aum"], 2) if stats["has_value"] else _NA
        ws2.append([name, stats["count"], aum_cell])

    for col in ws2.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"lookthrough_holdings_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
